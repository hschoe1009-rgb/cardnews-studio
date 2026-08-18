"""엑셀 DB. 저장할 때마다 cardnews_db.xlsx 를 갱신합니다.

시트 3개
  카드뉴스 : 편 단위 (날짜·주제·전략·캡션·경로)
  카드     : 카드 1장 단위 (내용·이미지 프롬프트)
  템플릿   : 사용된 템플릿/테마 통계
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config

SHEET_MAIN = "카드뉴스"
SHEET_CARDS = "카드"
SHEET_TEMPLATES = "템플릿"

HEADERS_MAIN = [
    "ID", "작성일", "수정일시", "주제", "좁힌 주제", "타깃",
    "컨셉축", "글포맷", "퍼널단계", "후킹", "CTA키워드",
    "카드수", "테마", "제목", "컨셉문장",
    "핵심키워드", "해시태그", "캡션", "댓글CTA", "다음편예고",
    "자가진단", "전략메모", "참고자료",
    "원문제목", "원문종류", "원문출처", "원문요약",
    "폴더경로", "상태", "메모",
]

HEADERS_CARDS = [
    "ID", "작성일", "주제", "카드번호", "슬러그", "역할", "템플릿",
    "배지", "타이틀", "서브", "본문", "노트", "팁", "이미지프롬프트", "이미지파일",
]

HEADERS_TEMPLATES = ["테마", "템플릿", "사용횟수", "최근사용일"]

HEADER_FILL = PatternFill("solid", fgColor="162033")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def _ensure_sheet(book: Workbook, name: str, headers: list[str]):
    if name in book.sheetnames:
        sheet = book[name]
        if sheet.max_row == 0 or sheet.cell(1, 1).value is None:
            sheet.append(headers)
            return sheet
        current = [c.value for c in sheet[1]][: len(headers)]
        if current == headers:
            return sheet
        # 열 구성이 바뀌었다. 기존 시트는 이름을 바꿔 보존하고 새로 만든다.
        stamp = datetime.now().strftime("%y%m%d%H%M%S")
        sheet.title = f"{name}_이전{stamp}"[:31]
    sheet = book.create_sheet(name)
    sheet.append(headers)
    return sheet


def _style_header(sheet, headers: list[str], widths: dict[int, int] | None = None) -> None:
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        letter = get_column_letter(col)
        sheet.column_dimensions[letter].width = (widths or {}).get(col, 16)
    sheet.freeze_panes = "A2"


def _open_book() -> Workbook:
    config.ensure_dirs()
    if config.DB_PATH.exists():
        try:
            return load_workbook(config.DB_PATH)
        except Exception:  # 손상 파일 → 백업 후 새로 만든다
            backup = config.DB_PATH.with_suffix(f".broken-{datetime.now():%y%m%d%H%M%S}.xlsx")
            config.DB_PATH.rename(backup)
    book = Workbook()
    book.remove(book.active)
    return book


def _delete_rows_by_id(sheet, project_id: str) -> None:
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row, 1).value == project_id:
            sheet.delete_rows(row)


def _join(values) -> str:
    if isinstance(values, (list, tuple)):
        return ", ".join(str(v) for v in values)
    return str(values or "")


def upsert(project: dict) -> Path:
    """프로젝트 1건을 DB에 반영(있으면 덮어씀)."""
    book = _open_book()
    main = _ensure_sheet(book, SHEET_MAIN, HEADERS_MAIN)
    cards_sheet = _ensure_sheet(book, SHEET_CARDS, HEADERS_CARDS)
    templates = _ensure_sheet(book, SHEET_TEMPLATES, HEADERS_TEMPLATES)

    pid = project["id"]
    _delete_rows_by_id(main, pid)
    _delete_rows_by_id(cards_sheet, pid)

    angle = project.get("angle") or {}
    src = project.get("source") or {}
    checks = project.get("self_check") or []
    check_text = " / ".join(
        f"{'O' if c.get('pass') else 'X'} {c.get('item', '')}" for c in checks
    )

    main.append([
        pid,
        project.get("created", ""),
        project.get("updated", ""),
        project.get("topic", ""),
        project.get("topic_refined", ""),
        project.get("audience", ""),
        angle.get("concept_axis", ""),
        angle.get("post_format", ""),
        angle.get("funnel_stage", ""),
        angle.get("hook", ""),
        angle.get("cta_keyword", ""),
        len(project.get("cards", [])),
        (project.get("theme") or {}).get("label", ""),
        project.get("title", ""),
        project.get("concept_sentence", ""),
        _join((project.get("keywords") or {}).get("core_keywords", [])),
        _join(project.get("hashtags", [])),
        project.get("caption", ""),
        project.get("cta_comment", ""),
        project.get("next_teaser", ""),
        check_text,
        project.get("strategy_notes", ""),
        _join(project.get("sources", [])),
        src.get("title", ""),
        src.get("kind", ""),
        src.get("origin", ""),
        src.get("summary", ""),
        str(config.PROJECTS_DIR / pid),
        project.get("status", ""),
        project.get("memo", ""),
    ])

    for card in project.get("cards", []):
        cards_sheet.append([
            pid,
            project.get("created", ""),
            project.get("topic", ""),
            card.get("index", ""),
            card.get("slug", ""),
            card.get("role", ""),
            card.get("template", ""),
            card.get("badge", ""),
            card.get("title", ""),
            card.get("subtitle", ""),
            card.get("body", ""),
            card.get("note", ""),
            card.get("tip", ""),
            card.get("image_prompt", ""),
            card.get("image", ""),
        ])

    _rebuild_templates(templates, main, cards_sheet)

    _style_header(main, HEADERS_MAIN, {4: 24, 5: 28, 18: 50, 22: 40, 24: 26, 27: 44, 28: 44})
    _style_header(cards_sheet, HEADERS_CARDS, {9: 24, 10: 26, 11: 30, 14: 50})
    _style_header(templates, HEADERS_TEMPLATES)
    for sheet in (main, cards_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    book.save(config.DB_PATH)
    return config.DB_PATH


def _rebuild_templates(sheet, main, cards_sheet) -> None:
    theme_by_id = {}
    date_by_id = {}
    for row in main.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        theme_by_id[row[0]] = row[12] or ""
        date_by_id[row[0]] = row[1] or ""

    counts: dict[tuple[str, str], list] = {}
    for row in cards_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        key = (theme_by_id.get(row[0], ""), row[6] or "")
        entry = counts.setdefault(key, [0, ""])
        entry[0] += 1
        used = date_by_id.get(row[0], "")
        if used > entry[1]:
            entry[1] = used

    sheet.delete_rows(2, max(sheet.max_row - 1, 0))
    for (theme, template), (count, last) in sorted(counts.items()):
        sheet.append([theme, template, count, last])


def remove(project_id: str) -> None:
    if not config.DB_PATH.exists():
        return
    book = load_workbook(config.DB_PATH)
    for name in (SHEET_MAIN, SHEET_CARDS):
        if name in book.sheetnames:
            _delete_rows_by_id(book[name], project_id)
    if SHEET_TEMPLATES in book.sheetnames and SHEET_MAIN in book.sheetnames:
        _rebuild_templates(book[SHEET_TEMPLATES], book[SHEET_MAIN], book[SHEET_CARDS])
    book.save(config.DB_PATH)
